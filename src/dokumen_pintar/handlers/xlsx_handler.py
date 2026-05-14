"""XLSX / XLSM format handler using openpyxl."""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from dokumen_pintar.errors import HandlerError, UnsupportedFormatError
from dokumen_pintar.handlers.base import (
    FormatHandler,
    HandlerCapability,
    default_registry,
)


def _load(path: Path, *, data_only: bool = False) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except (InvalidFileException, zipfile.BadZipFile) as exc:
        raise HandlerError(f"invalid xlsx file: {path} ({exc})") from exc
    except OSError as exc:
        raise HandlerError(f"cannot open xlsx file: {path} ({exc})") from exc


def _parse_sheet_ref(rest: str) -> tuple[str, str]:
    """Split '<sheet>!<a1>' into (sheet, a1)."""
    if "!" not in rest:
        raise HandlerError(f"expected '<sheet>!<ref>', got: {rest!r}")
    sheet, ref = rest.split("!", 1)
    if not sheet or not ref:
        raise HandlerError(f"invalid sheet reference: {rest!r}")
    return sheet, ref


def _require_sheet(wb: openpyxl.Workbook, name: str) -> Any:
    if name not in wb.sheetnames:
        raise HandlerError(f"sheet not found: {name!r}")
    return wb[name]


class XlsxHandler:
    """Handler for .xlsx / .xlsm spreadsheets via openpyxl."""

    name: str = "xlsx"
    extensions: tuple[str, ...] = (".xlsx", ".xlsm")
    capabilities: HandlerCapability = (
        HandlerCapability.READ_TEXT
        | HandlerCapability.STRUCTURED_GET
        | HandlerCapability.STRUCTURED_SET
        | HandlerCapability.STRUCTURED_DELETE
        | HandlerCapability.SEARCH_EXTRACTED
    )

    def detect(self, path: Path) -> bool:
        return path.suffix.lower() in self.extensions

    def read_meta(self, path: Path) -> dict[str, Any]:
        stat = path.stat()
        wb = _load(path)
        try:
            sheets: list[dict[str, Any]] = []
            for name in wb.sheetnames:
                ws = wb[name]
                sheets.append(
                    {
                        "name": name,
                        "rows": int(ws.max_row or 0),
                        "cols": int(ws.max_column or 0),
                    }
                )
            active_sheet = wb.active.title if wb.active is not None else None
        finally:
            wb.close()
        return {
            "format": self.name,
            "size": stat.st_size,
            "mtime": stat.st_mtime,
            "sheets": sheets,
            "active_sheet": active_sheet,
        }

    def read_text(self, path: Path, **_: Any) -> str:
        wb = _load(path)
        try:
            parts: list[str] = []
            for name in wb.sheetnames:
                ws = wb[name]
                lines: list[str] = [f"# Sheet: {name}"]
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join("" if v is None else str(v) for v in row))
                parts.append("\n".join(lines))
            return "\n".join(parts)
        finally:
            wb.close()

    def write_text(self, path: Path, content: str, **_: Any) -> None:
        raise UnsupportedFormatError("xlsx does not support write_text; use structured_set instead")

    def extract_for_search(self, path: Path) -> str:
        wb = _load(path, data_only=True)
        try:
            tokens: list[str] = []
            for name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows(values_only=True):
                    for v in row:
                        if v is None:
                            continue
                        tokens.append(str(v))
            return " ".join(tokens)
        finally:
            wb.close()

    def structured_get(self, path: Path, expr: str) -> Any:
        wb = _load(path)
        try:
            if expr == "sheets":
                return list(wb.sheetnames)

            if ":" not in expr:
                raise HandlerError(f"unsupported expression: {expr!r}")
            kind, rest = expr.split(":", 1)

            if kind == "sheet":
                ws = _require_sheet(wb, rest)
                return [list(row) for row in ws.iter_rows(values_only=True)]

            if kind == "cell":
                sheet, ref = _parse_sheet_ref(rest)
                ws = _require_sheet(wb, sheet)
                return ws[ref].value

            if kind == "range":
                sheet, ref = _parse_sheet_ref(rest)
                ws = _require_sheet(wb, sheet)
                cells = ws[ref]
                if not isinstance(cells, tuple):
                    return [[cells.value]]
                result: list[list[Any]] = []
                for row in cells:
                    if isinstance(row, tuple):
                        result.append([c.value for c in row])
                    else:  # pragma: no cover
                        result.append([row.value])
                return result

            raise HandlerError(f"unsupported expression: {expr!r}")
        finally:
            wb.close()

    def structured_set(self, path: Path, expr: str, value: Any) -> None:
        if ":" not in expr:
            raise HandlerError(f"unsupported expression: {expr!r}")
        kind, rest = expr.split(":", 1)
        wb = _load(path)
        try:
            if kind == "cell":
                sheet, ref = _parse_sheet_ref(rest)
                ws = _require_sheet(wb, sheet)
                ws[ref] = value

            elif kind == "range":
                sheet, ref = _parse_sheet_ref(rest)
                ws = _require_sheet(wb, sheet)
                if not isinstance(value, list):
                    raise HandlerError("range value must be a 2D list")
                cells = ws[ref]
                rows_iter: tuple[Any, ...]
                if isinstance(cells, tuple):
                    rows_iter = cells
                else:
                    rows_iter = (cells,)
                for row_cells, row_values in zip(rows_iter, value):
                    if not isinstance(row_values, list):
                        raise HandlerError("range value must be a 2D list")
                    if isinstance(row_cells, tuple):
                        target_cells = row_cells
                    else:
                        target_cells = (row_cells,)
                    for cell, cell_value in zip(target_cells, row_values):
                        cell.value = cell_value

            elif kind == "sheet":
                if not isinstance(value, list):
                    raise HandlerError("sheet value must be a 2D list")
                if rest in wb.sheetnames:
                    ws = wb[rest]
                    # Clear existing content.
                    if ws.max_row and ws.max_column:  # pragma: no branch
                        for row in ws.iter_rows(
                            min_row=1,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column,
                        ):
                            for cell in row:
                                cell.value = None
                else:
                    ws = wb.create_sheet(title=rest)
                for r_idx, row_values in enumerate(value, start=1):
                    if not isinstance(row_values, list):
                        raise HandlerError("sheet value must be a 2D list")
                    for c_idx, cell_value in enumerate(row_values, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=cell_value)

            else:
                raise HandlerError(f"unsupported expression: {expr!r}")

            try:
                wb.save(path)
            except OSError as exc:
                raise HandlerError(f"cannot save xlsx file: {path} ({exc})") from exc
        finally:
            wb.close()

    def structured_delete(self, path: Path, expr: str) -> None:
        if ":" not in expr:
            raise HandlerError(f"unsupported expression: {expr!r}")
        kind, rest = expr.split(":", 1)
        wb = _load(path)
        try:
            if kind == "sheet":
                if rest not in wb.sheetnames:
                    raise HandlerError(f"sheet not found: {rest!r}")
                del wb[rest]

            elif kind == "cell":
                sheet, ref = _parse_sheet_ref(rest)
                ws = _require_sheet(wb, sheet)
                ws[ref] = None

            elif kind == "row":
                sheet, ref = _parse_sheet_ref(rest)
                ws = _require_sheet(wb, sheet)
                try:
                    row_n = int(ref)
                except ValueError as exc:
                    raise HandlerError(f"invalid row number: {ref!r}") from exc
                if row_n < 1:
                    raise HandlerError(f"row must be >= 1, got {row_n}")
                ws.delete_rows(row_n, 1)

            elif kind == "col":
                sheet, ref = _parse_sheet_ref(rest)
                ws = _require_sheet(wb, sheet)
                try:
                    col_idx = column_index_from_string(ref)
                except ValueError as exc:
                    raise HandlerError(f"invalid column letter: {ref!r}") from exc
                ws.delete_cols(col_idx, 1)

            else:
                raise HandlerError(f"unsupported expression: {expr!r}")

            try:
                wb.save(path)
            except OSError as exc:
                raise HandlerError(f"cannot save xlsx file: {path} ({exc})") from exc
        finally:
            wb.close()


# Silence "imported but unused" for the helper re-export.
_ = get_column_letter

# Runtime-checkable protocol sanity assertion.
_handler: FormatHandler = XlsxHandler()
default_registry.register(_handler)
