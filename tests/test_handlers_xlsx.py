"""Tests for :class:`dokumen_pintar.handlers.xlsx_handler.XlsxHandler`."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from dokumen_pintar.errors import HandlerError, UnsupportedFormatError
from dokumen_pintar.handlers.xlsx_handler import XlsxHandler


@pytest.fixture
def handler() -> XlsxHandler:
    return XlsxHandler()


def _create_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["A3"] = "Bob"
    ws["B3"] = 25
    wb.save(path)


def test_detect(handler: XlsxHandler) -> None:
    assert handler.detect(Path("data.xlsx")) is True
    assert handler.detect(Path("data.xlsm")) is True
    assert handler.detect(Path("data.csv")) is False


def test_read_meta(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    meta = handler.read_meta(target)
    assert meta["format"] == "xlsx"
    assert len(meta["sheets"]) >= 1
    assert meta["sheets"][0]["name"] == "Data"
    assert meta["sheets"][0]["rows"] >= 3
    assert meta["sheets"][0]["cols"] >= 2


def test_read_text(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    text = handler.read_text(target)
    assert "Alice" in text
    assert "Bob" in text


def test_write_text_raises(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(UnsupportedFormatError):
        handler.write_text(target, "content")


def test_extract_for_search(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    text = handler.extract_for_search(target)
    assert "Alice" in text
    assert "Bob" in text


def test_structured_get_sheets(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    sheets = handler.structured_get(target, "sheets")
    assert isinstance(sheets, list)
    assert "Data" in sheets


def test_structured_get_sheet(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    data = handler.structured_get(target, "sheet:Data")
    assert isinstance(data, list)
    assert data[0] == ["Name", "Age"]


def test_structured_get_cell(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    val = handler.structured_get(target, "cell:Data!A2")
    assert val == "Alice"


def test_structured_get_range(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    data = handler.structured_get(target, "range:Data!A1:B2")
    assert isinstance(data, list)
    assert len(data) == 2


def test_structured_get_unsupported(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="unsupported"):
        handler.structured_get(target, "invalid_expr")


def test_structured_set_cell(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    handler.structured_set(target, "cell:Data!A2", "Charlie")
    val = handler.structured_get(target, "cell:Data!A2")
    assert val == "Charlie"


def test_structured_set_unsupported(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="unsupported"):
        handler.structured_set(target, "invalid_expr", "val")


def test_structured_delete_cell(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    handler.structured_delete(target, "cell:Data!A2")
    val = handler.structured_get(target, "cell:Data!A2")
    assert val is None


def test_structured_delete_row(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    handler.structured_delete(target, "row:Data!2")
    val = handler.structured_get(target, "cell:Data!A2")
    # After deleting row 2, row 3 (Bob) shifts up
    assert val == "Bob"


def test_structured_delete_sheet(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Keep"
    ws2 = wb.create_sheet("Remove")
    ws1["A1"] = "kept"
    ws2["A1"] = "removed"
    wb.save(target)

    handler.structured_delete(target, "sheet:Remove")
    sheets = handler.structured_get(target, "sheets")
    assert "Remove" not in sheets
    assert "Keep" in sheets


def test_structured_delete_unsupported(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="unsupported"):
        handler.structured_delete(target, "invalid_expr")


def test_structured_delete_sheet_not_found(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="not found"):
        handler.structured_delete(target, "sheet:NonExistent")


# ── Additional XLSX coverage ──


def test_structured_set_range(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    handler.structured_set(target, "range:Data!A1:B2", [["X", "Y"], ["1", "2"]])
    val = handler.structured_get(target, "cell:Data!A1")
    assert val == "X"


def test_structured_set_sheet_new(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    handler.structured_set(target, "sheet:NewSheet", [["h1", "h2"], ["v1", "v2"]])
    sheets = handler.structured_get(target, "sheets")
    assert "NewSheet" in sheets


def test_structured_set_sheet_overwrite(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    handler.structured_set(target, "sheet:Data", [["new1"], ["new2"]])
    val = handler.structured_get(target, "cell:Data!A1")
    assert val == "new1"


def test_structured_delete_col(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    handler.structured_delete(target, "col:Data!A")
    val = handler.structured_get(target, "cell:Data!A1")
    assert val == "Age"  # Column B shifted to A


def test_structured_get_range_single_cell(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    result = handler.structured_get(target, "range:Data!A1:A1")
    assert isinstance(result, list)


def test_structured_set_non_list_sheet_raises(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="2D list"):
        handler.structured_set(target, "sheet:Data", "not a list")


def test_structured_set_non_list_range_raises(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="2D list"):
        handler.structured_set(target, "range:Data!A1:B2", "not a list")


def test_structured_get_sheet_not_found(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="not found"):
        handler.structured_get(target, "sheet:NoSheet")


# ── More XLSX coverage ──


def test_invalid_xlsx_file(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "bad.xlsx"
    target.write_bytes(b"not a zip file at all")
    with pytest.raises(HandlerError, match="invalid xlsx"):
        handler.read_meta(target)


def test_structured_get_unsupported_kind(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="unsupported"):
        handler.structured_get(target, "badkind:Data")


def test_structured_set_unsupported_kind(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="unsupported"):
        handler.structured_set(target, "badkind:Data", "val")


def test_structured_delete_unsupported_kind(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="unsupported"):
        handler.structured_delete(target, "badkind:Data")


def test_structured_set_range_non_2d_list_row(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="2D list"):
        handler.structured_set(target, "range:Data!A1:B2", [["ok", "ok"], "not a list"])


def test_structured_set_sheet_non_2d_list_row(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="2D list"):
        handler.structured_set(target, "sheet:Data", [["ok"], "not a list"])


def test_structured_delete_row_invalid(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="invalid row"):
        handler.structured_delete(target, "row:Data!abc")


def test_structured_delete_row_negative(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="row must be >= 1"):
        handler.structured_delete(target, "row:Data!0")


def test_structured_delete_col_invalid(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="invalid column"):
        handler.structured_delete(target, "col:Data!999")


def test_parse_sheet_ref_no_excl(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="expected"):
        handler.structured_get(target, "cell:DataA1")


def test_parse_sheet_ref_empty_parts(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    with pytest.raises(HandlerError, match="invalid sheet"):
        handler.structured_get(target, "cell:!A1")


def test_structured_get_range_single(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    result = handler.structured_get(target, "range:Data!A1")
    assert isinstance(result, list)
    assert result == [["Name"]]


def test_active_sheet_meta(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "test.xlsx"
    _create_xlsx(target)
    meta = handler.read_meta(target)
    assert meta["active_sheet"] == "Data"


def test_load_oserror(handler: XlsxHandler, tmp_path: Path) -> None:
    from unittest.mock import patch
    target = tmp_path / "oserr.xlsx"
    # Create a valid xlsx, then patch load to raise OSError
    _create_xlsx(target)
    with patch("openpyxl.load_workbook", side_effect=OSError("disk error")):
        with pytest.raises(HandlerError, match="cannot open xlsx"):
            handler.read_meta(target)


def test_extract_for_search_skips_none(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "sparse.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "hello"
    ws["B1"] = None
    ws["C1"] = "world"
    wb.save(target)
    wb.close()
    text = handler.extract_for_search(target)
    assert "hello" in text
    assert "world" in text
    assert "None" not in text


def test_structured_get_range_non_tuple_row(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "rng.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "val"
    wb.save(target)
    wb.close()
    # range:S!A1 should return single cell as [[val]]
    result = handler.structured_get(target, "range:S!A1")
    assert result == [["val"]]


def test_structured_set_range_single_cell(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "rngset.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "old"
    wb.save(target)
    wb.close()
    handler.structured_set(target, "range:S!A1", [["new"]])
    result = handler.structured_get(target, "cell:S!A1")
    assert result == "new" or result == [["new"]]


def test_structured_set_save_oserror(handler: XlsxHandler, tmp_path: Path) -> None:
    from unittest.mock import patch
    target = tmp_path / "saveerr.xlsx"
    _create_xlsx(target)
    with patch("openpyxl.Workbook.save", side_effect=OSError("disk full")):
        with pytest.raises(HandlerError, match="cannot save xlsx"):
            handler.structured_set(target, "cell:Data!A1", "X")


def test_structured_delete_save_oserror(handler: XlsxHandler, tmp_path: Path) -> None:
    from unittest.mock import patch
    target = tmp_path / "delsave.xlsx"
    _create_xlsx(target)
    with patch("openpyxl.Workbook.save", side_effect=OSError("disk full")):
        with pytest.raises(HandlerError, match="cannot save xlsx"):
            handler.structured_delete(target, "cell:Data!A1")


def test_structured_get_range_single_col_multi_row(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "sc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"  # type: ignore[union-attr]
    ws["A1"] = "r1"  # type: ignore[index]
    ws["A2"] = "r2"  # type: ignore[index]
    ws["A3"] = "r3"  # type: ignore[index]
    wb.save(str(target))
    # A1:A3 single-column range: openpyxl returns tuple of Cell (not tuple of tuple)
    result = handler.structured_get(target, "range:Sheet!A1:A3")
    assert isinstance(result, list)
    assert any("r1" in str(r) for r in result)


def test_structured_set_sheet_empty_existing(handler: XlsxHandler, tmp_path: Path) -> None:
    target = tmp_path / "emptysh.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"  # type: ignore[union-attr]
    # Leave sheet empty (max_row/max_column will be 0 or None)
    wb.save(str(target))
    handler.structured_set(target, "sheet:Data", [["a", "b"], ["1", "2"]])
    result = handler.structured_get(target, "cell:Data!A1")
    assert result == "a"
